"""
- openpyxl é uma biblioteca Python para ler / gravar arquivos xlsx / xlsm /
xltx / xltm do Excel 2010.
- Por padrão, o openpyxl não protege contra explosão quadrática ou ataques
xml de bilhões de risos. Para se proteger contra esses ataques, instale o
defusedxml.
- https://openpyxl.readthedocs.io/en/stable/
- pip install openpyxl
- pipenv install openpyxl
"""
from random import uniform

import openpyxl

planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = f'Cód_{numero_pedido}'
    planilha1.cell(linha, 2).value = f'{1200 + linha}'

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = f'R$ {preco}'

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'cód {linha + round(uniform(10, 99))}'
    planilha2.cell(linha, 2).value = f'R$ {round(uniform(10, 99), 2)}'
    planilha2.cell(linha, 3).value = f'R$ {round(uniform(10, 99), 2)}'
    planilha2.cell(linha, 4).value = f'R$ {round(uniform(10, 99), 2)}'

print()
for linha in planilha1:
    if linha[0].value is not None:
        print(linha[0].value, linha[1].value, linha[2].value)

print()
for linha in planilha2:
    if linha[0].value is not None:
        print(linha[0].value, linha[1].value, linha[2].value)

planilha.save('nova_planilha.xlsx')
