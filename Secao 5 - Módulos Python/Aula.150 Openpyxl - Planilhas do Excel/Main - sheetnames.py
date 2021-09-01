"""
- https://openpyxl.readthedocs.io/en/stable/
- pip install openpyxl
- pipenv install openpyxl
"""
import openpyxl

casos_covid19 = openpyxl.load_workbook('Casos Diários de Covid-19.xlsx')
casos_escalados = openpyxl.load_workbook('Casos Escalados de Covid-19.xlsx')

nome_planilhas_01 = casos_covid19.sheetnames
nome_planilhas_02 = casos_escalados.sheetnames

print()
print(f'Nome da(s) Planilha(s): {nome_planilhas_01}')
print(f'Nome da(s) Planilha(s): {nome_planilhas_02}')

planilha_01 = casos_covid19['Casos Diários na Cidade']
planilha_02 = casos_escalados['Casos Escalados na Cidade']

print()
print(f"Valor da célula C4 e D4: {planilha_01['c4'].value}: {planilha_01['d4'].value}")
print(f"Valor da célula D1 e D94: {planilha_02['d1'].value}: {planilha_02['d94'].value}")

print()
for campo in planilha_01['c']:
    print(f'Coluna C: {campo.value}')

print()
for linha in planilha_01['c1:d10']:
    for coluna in linha:
        print(coluna.value)

print()
for linha in planilha_01:
    if linha[0].value is not None:
        print(linha[0].value, linha[1].value, linha[2].value, linha[3].value)

print()
for linha in planilha_02:
    if linha[0].value is not None:
        print(linha[0].value, linha[1].value, linha[2].value, linha[3].value)
